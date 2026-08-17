import re
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010\013\014\016-\037]')

def clean_val(v: Any) -> str:
    if v is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(v))


def _canonicalize_lookup_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _build_preview_lookup(preview_sections: List[Dict[str, Any]]) -> Dict[str, Any]:
    lookup: Dict[str, Any] = {}
    for sec in preview_sections:
        for field in sec.get("fields", []):
            value = field.get("value")
            if value in (None, ""):
                continue
            for candidate in (
                field.get("field_name"),
                field.get("label"),
                field.get("id"),
                _canonicalize_lookup_key(field.get("field_name")),
                _canonicalize_lookup_key(field.get("label")),
            ):
                if candidate:
                    lookup[str(candidate).strip().lower()] = value
    return lookup


def _preview_aliases_for(display_key: str) -> List[str]:
    alias_map = {
        "tender_value_display": ["tender_value_gst_inclusive", "tender_value_gst", "tender_value", "estimated_value", "tender value (gst inclusive)", "tender value"],
        "tender_fee_amount_display": ["tender_fee_amount", "tender fee"],
        "processing_fee_amount_display": ["processing_fee_amount", "processing fee amount"],
        "emd_amount_display": ["emd_amount", "emd amount", "emd"],
        "bid_due_date_time": ["bid_submission_deadline", "bid due date and time", "bid submission deadline"],
        "payment_terms_supply_display": ["payment_terms_supply", "payment terms supply", "payment terms %"],
        "payment_terms_installation_display": ["payment_terms_installation", "payment terms installation", "payment terms installation (%)"],
        "delivery_time_supply_display": ["delivery_time_supply", "delivery time supply", "delivery time supply (days)"],
        "pbg_mode_display": ["pbg_mode", "pbg mode"],
        "pbg_percentage_display": ["pbg_percentage", "pbg percentage"],
        "pbg_duration_display": ["pbg_duration", "pbg duration (months)"],
        "sd_mode_display": ["sd_mode", "security deposit mode"],
        "sd_percentage_display": ["sd_percentage", "security deposit %"],
        "sd_duration_display": ["sd_duration", "sd duration (months)"],
        "ld_percentage_display": ["ld_percentage_per_week", "ld percentage per week"],
        "max_ld_percentage_display": ["max_ld_percentage", "max ld percentage"],
        "experience_years_display": ["eligibility_criterion_years", "experience years"],
        "bid_validity_days_display": ["bid_validity_period", "bid validity period"],
        "custom_eligibility_criteria_display": ["custom_eligibility_criteria", "custom eligibility criteria"],
        "avg_annual_turnover_value_display": ["avg_annual_turnover_value", "annual avg turnover", "annual_avg_turnover_value"],
        "working_capital_value_display": ["working_capital_value", "working capital"],
        "solvency_certificate_value_display": ["solvency_certificate_value", "solvency certificate"],
        "courier_address_display": ["courier_address", "courier address"],
        "client_name_1_display": ["client_name_1", "client contacts", "client contact person"],
        "client_name_2_display": ["client_name_2", "client contacts 2", "client contacts ii"],
        "client_name_3_display": ["client_name_3", "client contacts 3", "client contacts iii"],
    }
    return alias_map.get(display_key, [])

from backend.app.services.csv_schema import (
    BIDDER_READINESS_SUMMARY_LAYOUT,
    INFOSHEET_PAGE1_LAYOUT,
    INFOSHEET_PAGE2_LAYOUT,
    INFOSHEET_COLUMN_WIDTHS,
    INFOSHEET_DATA_KEYS,
)

def apply_cell_style(
    cell: Any,
    style_name: str,
    cell_def: Dict[str, Any],
    is_atc_override: bool = False,
    field_status: Optional[str] = None
) -> None:
    # Base font
    font_name = "Segoe UI"
    font_size = 10
    font_color = "000000"
    bold = cell_def.get("bold", False)
    
    # Fills & Status visual treatments
    fill = None
    if field_status == "MISSING" and cell_def.get("kind") == "value":
        fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Soft Red fill
        font_color = "78281F" # Dark Red text
        bold = True
    elif field_status == "OK_FALLBACK" and cell_def.get("kind") == "value":
        fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Soft Yellow fill
        font_color = "7D6608" # Dark Brown text
        bold = True
    elif field_status == "NOT_APPLICABLE" and cell_def.get("kind") == "value":
        fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid") # Soft Gray fill
        font_color = "566573" # Slate Gray text
        bold = False
    elif is_atc_override:
        fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        font_color = "0E6251"
        bold = True
    elif style_name == "section_header":
        fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        font_color = "FFFFFF"
        font_size = 11
        bold = True
    elif style_name == "subsection_header":
        fill = PatternFill(start_color="E5E8E8", end_color="E5E8E8", fill_type="solid")
        font_color = "2C3E50"
        bold = True
    elif style_name == "label_pink":
        fill = PatternFill(start_color="FFF0F2", end_color="FFF0F2", fill_type="solid")
        font_color = "78281F"
    elif style_name == "value_pink":
        fill = PatternFill(start_color="FFF0F2", end_color="FFF0F2", fill_type="solid")
        font_color = "000000"
    elif style_name == "label_yellow":
        fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        font_color = "7D6608"
    elif style_name == "value_yellow":
        fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        font_color = "000000"
    elif style_name == "value_blue":
        fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        font_color = "1B4F72"
    
    cell.font = Font(name=font_name, size=font_size, bold=bold, color=font_color)
    if fill:
        cell.fill = fill
        
    # Alignments
    horizontal_align = cell_def.get("align", "left")
    if cell_def.get("kind") == "header":
        horizontal_align = "center"
    cell.alignment = Alignment(
        horizontal=horizontal_align, 
        vertical="center", 
        wrap_text=cell_def.get("wrap", True)
    )
    
    # Border
    thin_side = Side(border_style="thin", color="CCCCCC")
    cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def render_layout(
    ws: Any,
    layout: List[Dict[str, Any]],
    data: Dict[str, str],
    field_sources: Dict[str, str] = None,
    field_statuses: Dict[str, str] = None,
    start_row: int = 1
) -> int:
    field_sources = field_sources or {}
    field_statuses = field_statuses or {}
    current_row = start_row
    for row_def in layout:
        ws.row_dimensions[current_row].height = row_def.get("height", 20)
        c_idx = 1
        for cell_def in row_def.get("cells", []):
            colspan = cell_def.get("colspan", 1)
            # Merge if colspan > 1
            if colspan > 1:
                ws.merge_cells(start_row=current_row, start_column=c_idx, end_row=current_row, end_column=c_idx + colspan - 1)
            
            # Resolve value
            val = ""
            key = None
            f_status = None
            if cell_def["kind"] in ("label", "header"):
                val = cell_def.get("text") or ""
            elif cell_def["kind"] == "value":
                key = cell_def.get("key")
                val = data.get(key) if key in data else "N/A"
                if val is None or val == "":
                    val = "N/A"
                f_status = field_statuses.get(key)
            
            # Write to top-left cell
            ws.cell(row=current_row, column=c_idx, value=clean_val(val))
            
            # Check if this cell is an ATC override
            is_atc = (cell_def["kind"] == "value" and key and field_sources.get(key) == "atc")
            
            # Apply styling to all cells in the merged range
            for col in range(c_idx, c_idx + colspan):
                cell = ws.cell(row=current_row, column=col)
                style_name = cell_def.get("style", "plain")
                apply_cell_style(cell, style_name, cell_def, is_atc_override=is_atc, field_status=f_status)
                
            c_idx += colspan
        current_row += 1
    return current_row


def render_flat_sections_sheet(wb: Workbook, sections: List[Dict[str, Any]], title: str = "Preview Fields") -> None:
    ws = wb.create_sheet(title=title)
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Row Number",
        "Field Section",
        "Field Name",
        "Preview Value",
        "Confidence",
        "Status",
        "Document Source",
        "Source Snippet",
    ]
    ws.append(headers)

    row_num = 1
    for sec in sections:
        for field in sec.get("fields", []):
            src_tag = field.get("source") or "MAIN"
            if src_tag in ("main_tender", "MAIN"):
                src_str = "MAIN"
            elif src_tag in ("atc", "ATC"):
                src_str = "ATC"
            elif src_tag == "ambiguous_preserved":
                src_str = "AMBIGUOUS_PRESERVED"
            elif src_tag == "derived":
                src_str = "DERIVED"
            else:
                src_str = str(src_tag).upper()

            ws.append([
                row_num,
                clean_val(sec.get("title", "")),
                clean_val(field.get("label", "")),
                clean_val(field.get("value", "")),
                clean_val(f"{field.get('confidence', 0)}%"),
                clean_val(field.get("status", "extracted")),
                clean_val(src_str),
                clean_val(field.get("sourceSnippet", "")),
            ])
            row_num += 1

    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10)
    thin_side = Side(border_style="thin", color="CCCCCC")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    for r_idx in range(2, row_num + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = cell_font
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 5, 6, 7) else "left",
                vertical="center",
                wrap_text=True,
            )

    ws.row_dimensions[1].height = 28
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


def generate_info_sheet_csv(data: Any, output_path: str) -> None:
    """
    Writes extracted fields into a standard XLSX sheet format using openpyxl.
    Supports rendering a visual layout dict or a list-of-sections flat format.
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    if isinstance(data, dict):
        preview_sections = data.pop("_info_sheet_sections", None)
        field_sources = data.pop("_info_sheet_sources", {})
        status_summary = data.get("status_summary", {})
        missing_fields = data.get("missing_fields", [])
        field_statuses = data.get("_info_sheet_statuses", {})
        
        # 1:1 key mapping validation
        data_copy = dict(data)
        preview_lookup = _build_preview_lookup(preview_sections) if isinstance(preview_sections, list) else {}
        missing_keys = set(INFOSHEET_DATA_KEYS) - set(data_copy.keys())
        extra_keys = set(data_copy.keys()) - set(INFOSHEET_DATA_KEYS)

        for key in list(missing_keys):
            if data_copy.get(key) not in (None, ""):
                continue
            for alias in _preview_aliases_for(key):
                alias_value = preview_lookup.get(alias.lower())
                if alias_value not in (None, ""):
                    data_copy[key] = alias_value
                    break
        
        for k in missing_keys:
            if data_copy.get(k) in (None, ""):
                data_copy[k] = "N/A"
        for k in extra_keys:
            data_copy.pop(k, None)
            
        data = data_copy

        ws = wb.create_sheet(title="InfoSheet")
        ws.views.sheetView[0].showGridLines = True
        # Set column widths
        for col_idx, width in enumerate(INFOSHEET_COLUMN_WIDTHS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
        # Render Top Health & Completeness Summary Header Block (Rows 1-3)
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 24
        c1 = ws.cell(row=1, column=1, value="TENDER EXTRACTION COMPLETENESS & HEALTH SUMMARY")
        c1.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        c1.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        c1.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:F2")
        ws.row_dimensions[2].height = 22
        ok_cnt = status_summary.get("OK", 0)
        fb_cnt = status_summary.get("OK_FALLBACK", 0)
        na_cnt = status_summary.get("NOT_APPLICABLE", 0)
        ms_cnt = status_summary.get("MISSING", 0)
        tot_cnt = ok_cnt + fb_cnt + na_cnt + ms_cnt
        
        sum_str = f"Total Schema Fields: {tot_cnt}  |  OK (Extracted): {ok_cnt}  |  OK (Fallback): {fb_cnt}  |  N/A (Not Applicable): {na_cnt}  |  ⚠️ MISSING: {ms_cnt}"
        c2 = ws.cell(row=2, column=1, value=sum_str)
        c2.fill = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
        c2.font = Font(name="Segoe UI", size=10, bold=True, color="2C3E50")
        c2.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A3:F3")
        ws.row_dimensions[3].height = 20
        if missing_fields:
            clean_missing_names = [m.replace("_display", "").replace("_", " ").title() for m in missing_fields[:8]]
            msg_missing = f"⚠️ ACTION REQUIRED — Missing Fields ({len(missing_fields)}): " + ", ".join(clean_missing_names)
            if len(missing_fields) > 8:
                msg_missing += f" (+{len(missing_fields)-8} more)"
            c3 = ws.cell(row=3, column=1, value=msg_missing)
            c3.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            c3.font = Font(name="Segoe UI", size=9, bold=True, color="78281F")
        else:
            c3 = ws.cell(row=3, column=1, value="✅ COMPLETE — All expected fields successfully resolved with zero missing extractions.")
            c3.fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
            c3.font = Font(name="Segoe UI", size=9, bold=True, color="145A32")
        c3.alignment = Alignment(horizontal="center", vertical="center")

        # Row 4 spacer
        ws.row_dimensions[4].height = 12

        # Render Bidder Readiness & Qualification Summary Block
        next_row = render_layout(ws, BIDDER_READINESS_SUMMARY_LAYOUT, data, field_sources=field_sources, field_statuses=field_statuses, start_row=5)

        # Spacer row between Readiness Summary and Detailed Fields
        ws.row_dimensions[next_row].height = 15
        next_row += 1

        next_row = render_layout(ws, INFOSHEET_PAGE1_LAYOUT, data, field_sources=field_sources, field_statuses=field_statuses, start_row=next_row)
        
        # Spacer row between pages
        ws.row_dimensions[next_row].height = 15
        next_row += 1
        
        render_layout(ws, INFOSHEET_PAGE2_LAYOUT, data, field_sources=field_sources, field_statuses=field_statuses, start_row=next_row)

        if isinstance(preview_sections, list) and preview_sections:
            render_flat_sections_sheet(wb, preview_sections)
        
    else:
        # Fallback to the old list-of-sections flat format
        render_flat_sections_sheet(wb, data, title="InfoSheet")
            
    wb.save(output_path)

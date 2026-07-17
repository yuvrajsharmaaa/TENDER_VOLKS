import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from pathlib import Path

new_path = Path('backend/app/storage/jobs/live-regen-2d3467f4/GAIL VRLA Jamnagar_InfoSheet.xlsx')
old_path = Path('backend/app/storage/jobs/35580348-246b-49f7-86a0-175c1bfd64ca/GAIL VRLA Jamnagar_InfoSheet.xlsx')

def read_all_sheets(p):
    wb = openpyxl.load_workbook(p, data_only=True)
    print('Sheets in', p.name, ':', wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f'  --- Sheet: {sheet} ---')
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                safe_row = tuple(str(c).encode('ascii', 'replace').decode('ascii') if c is not None else None for c in row)
                print('  ', safe_row)

print('=== OLD ===')
read_all_sheets(old_path)
print()
print('=== NEW ===')
read_all_sheets(new_path)

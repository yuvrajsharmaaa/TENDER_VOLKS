import pytest
from backend.app.services.tender_mapper import (
    build_infosheet_data,
    FIELD_STATUS_OK,
    FIELD_STATUS_OK_FALLBACK,
    FIELD_STATUS_NOT_APPLICABLE,
    FIELD_STATUS_MISSING
)
from backend.app.services.info_sheet_generator import generate_info_sheet_csv
from openpyxl import load_workbook
import tempfile
from pathlib import Path

def test_field_status_classification():
    mock_sections = [
        {
            "title": "Tender Information",
            "fields": [
                {"id": "f-emd_amount", "label": "EMD Amount", "value": 50000.0, "status": "extracted", "confidence": 95.0, "source": "main_tender"},
                {"id": "f-pbg_percentage", "label": "PBG Percentage", "value": 5.0, "status": "extracted", "confidence": 95.0, "source": "main_tender"},
                {"id": "f-sd_percentage", "label": "Security Deposit %", "value": None, "status": "missing", "confidence": 0.0, "source": "main_tender"}
            ]
        }
    ]
    
    mock_pages = [{"page": 1, "text": "EMD Amount: 50000. PBG: 5%. No Security Deposit specified."}]
    
    data = build_infosheet_data(mock_sections, mock_pages, job_id="test-status-job")
    
    status_summary = data.get("status_summary", {})
    missing_fields = data.get("missing_fields", [])
    field_statuses = data.get("_info_sheet_statuses", {})
    
    assert FIELD_STATUS_OK in status_summary
    assert FIELD_STATUS_NOT_APPLICABLE in status_summary
    assert FIELD_STATUS_MISSING in status_summary
    
    # SD fields should be NOT_APPLICABLE because PBG is 5% and SD percentage is missing
    assert field_statuses.get("sd_mode_display") == FIELD_STATUS_NOT_APPLICABLE
    assert field_statuses.get("sd_percentage_display") == FIELD_STATUS_NOT_APPLICABLE
    
    # EMD Amount should be OK
    assert field_statuses.get("emd_amount_display") == FIELD_STATUS_OK
    
    # Check JSON export fields
    assert isinstance(missing_fields, list)

def test_infosheet_excel_rendering():
    mock_data = {
        "organization": "Green Gas Limited",
        "tender_name": "UPS Batteries Procurement",
        "tender_id_display": "GEM/2026/B/7772525",
        "emd_amount_display": "50000.0",
        "pbg_percentage_display": "5.0%",
        "sd_mode_display": "N/A",
        "courier_address_display": "⚠️ MISSING",
        "status_summary": {
            FIELD_STATUS_OK: 10,
            FIELD_STATUS_OK_FALLBACK: 2,
            FIELD_STATUS_NOT_APPLICABLE: 5,
            FIELD_STATUS_MISSING: 1
        },
        "missing_fields": ["courier_address_display"],
        "_info_sheet_statuses": {
            "organization": FIELD_STATUS_OK,
            "tender_name": FIELD_STATUS_OK,
            "tender_id_display": FIELD_STATUS_OK,
            "emd_amount_display": FIELD_STATUS_OK,
            "pbg_percentage_display": FIELD_STATUS_OK,
            "sd_mode_display": FIELD_STATUS_NOT_APPLICABLE,
            "courier_address_display": FIELD_STATUS_MISSING
        }
    }
    
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = str(Path(tmpdir) / "test_InfoSheet.xlsx")
        generate_info_sheet_csv(mock_data, out_path)
        
        wb = load_workbook(out_path)
        assert "InfoSheet" in wb.sheetnames
        ws = wb["InfoSheet"]
        
        # Check header summary row 1 & 2
        assert "TENDER EXTRACTION COMPLETENESS & HEALTH SUMMARY" in str(ws["A1"].value)
        assert "Total Schema Fields" in str(ws["A2"].value)
        assert "⚠️ ACTION REQUIRED — Missing Fields" in str(ws["A3"].value)

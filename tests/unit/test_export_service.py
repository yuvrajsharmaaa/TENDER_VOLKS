import os
import csv
import pytest
from pathlib import Path
from backend.app.services.export_service import (
    export_tender_information_csv,
    export_tender_evidence_csv,
    export_page_aware_tender_sheets
)

@pytest.fixture
def temp_output_dir(tmp_path):
    return str(tmp_path)

def test_export_tender_information_csv(temp_output_dir):
    summary_data = {
        "tender_id": 12345,
        "tender_value": 2500000.0,
        "emd_required": "Yes",
        "emd_mode": ["BG", "DD"],
        "te_recommendation": None
    }
    
    filepath = export_tender_information_csv(summary_data, temp_output_dir)
    assert os.path.exists(filepath)
    assert "tender_12345_export.csv" in filepath
    
    # Verify content
    with open(filepath, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["tender_id"] == "12345"
        assert rows[0]["tender_value"] == "2500000.0"
        assert rows[0]["emd_mode"] == "BG|DD"
        assert rows[0]["te_recommendation"] == ""

def test_export_tender_evidence_csv(temp_output_dir):
    evidence_data = [
        {
            "tender_id": 12345,
            "field_name": "tender_value",
            "extracted_value_raw": "Rs 25 Lakh",
            "normalized_value": 2500000.0,
            "page_number": 1,
            "confidence": 0.9,
            "text_snippet": "Estimated cost Rs 25 Lakh",
            "extraction_timestamp": "2026-07-07 12:00:00"
        }
    ]
    
    filepath = export_tender_evidence_csv(evidence_data, 12345, temp_output_dir)
    assert os.path.exists(filepath)
    assert "tender_12345_evidence.csv" in filepath
    
    # Verify content
    with open(filepath, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["tender_id"] == "12345"
        assert rows[0]["field_name"] == "tender_value"
        assert rows[0]["page_number"] == "1"

def test_export_page_aware_tender_sheets(temp_output_dir):
    summary_data = {"tender_id": 12345, "tender_value": 2500000.0}
    evidence_data = [{"tender_id": 12345, "field_name": "tender_value", "page_number": 1}]
    
    summary_path, evidence_path = export_page_aware_tender_sheets(summary_data, evidence_data, temp_output_dir)
    assert os.path.exists(summary_path)
    assert os.path.exists(evidence_path)


def test_generate_info_sheet_with_atc_sources(temp_output_dir):
    from backend.app.services.info_sheet_generator import generate_info_sheet_csv
    from openpyxl import load_workbook

    data = {
        "bid_validity_days_display": "120 Days",
        "_info_sheet_sources": {
            "bid_validity_days_display": "atc"
        },
        "_info_sheet_sections": [
            {
                "title": "General Info",
                "fields": [
                    {
                        "label": "Bid Validity",
                        "value": "120 Days",
                        "confidence": 95,
                        "status": "extracted",
                        "source": "atc",
                        "sourceSnippet": "Bid Offer Validity: 120 Days"
                    }
                ]
            }
        ]
    }
    
    output_path = os.path.join(temp_output_dir, "test_info_sheet.xlsx")
    generate_info_sheet_csv(data, output_path)
    
    assert os.path.exists(output_path)
    wb = load_workbook(output_path)
    assert "InfoSheet" in wb.sheetnames
    assert "Preview Fields" in wb.sheetnames

    # Check Preview Fields headers for Document Source
    ws_preview = wb["Preview Fields"]
    headers = [cell.value for cell in ws_preview[1]]
    assert "Document Source" in headers
    
    # Check data row in Preview Fields
    row2_vals = [cell.value for cell in ws_preview[2]]
    assert "ATC" in row2_vals

    wb.close()

